#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Exporta Monitoreo_Creativos_Meta.xlsx (ya lleno con fill_monitoreo.py y RECALCULADO
con recalc.py de la skill xlsx) a un data.json liviano para el dashboard estático
(index.html en GitHub Pages).

Uso:
  python export_dashboard.py "<ruta>/Monitoreo_Creativos_Meta.xlsx" data.json
  python export_dashboard.py "<ruta>/Monitoreo_Creativos_Meta.xlsx" data.json --fecha "17 jul 2026"

Requisitos:
- El archivo YA debe estar recalculado (data_only=True necesita los valores en caché;
  si no corriste recalc.py después de fill_monitoreo.py, las columnas de fórmula
  como CPI, ESTADO, Acción saldrán vacías en el JSON).
- No modifica el xlsx. Solo lee.
"""
import sys
import json
import argparse
from datetime import date, datetime
import openpyxl

SHEET_NAME = "Monitoreo Piezas"
HEADER_ROW = 4
FIRST_DATA_ROW = 5

# Columnas que nos interesan y su clave en el JSON. Deben calzar EXACTO con los
# encabezados reales del archivo (fila 4 de 'Monitoreo Piezas'). Verificado contra
# Monitoreo_Creativos_Meta.xlsx real: los nombres incluyen unidades entre paréntesis
# en varias columnas (CPI ($), CPI ref ($), Hook rate, Acción recomendada).
COLS = {
    "Pieza ID": "pieza_id",
    "Ruta": "ruta",
    "Conjunto": "conjunto",
    "Formato": "formato",
    "Ángulo / Tema": "angulo",
    "Fecha inicio test": "fecha_inicio",
    "Cohorte": "cohorte",
    "Días en test": "dias_en_test",
    "Últ. actividad": "ultima_actividad",
    "Estado anuncio": "estado_anuncio",
    "Gasto ($)": "gasto",
    "Impresiones": "impresiones",
    "Clics": "clics",
    "Installs": "installs",
    "Views 3s": "views3s",
    "KYC compl. (opc.)": "kyc",
    "CPI ($)": "cpi",
    "CTR": "ctr",
    "Hook rate": "hook",
    "KYC/inst": "kyc_por_install",
    "CPI ref ($)": "cpi_ref",
    "CPI vs ref": "cpi_vs_ref",
    "ESTADO": "estado",
    "Acción recomendada": "accion",
    "Notas": "notas",
    "Base CPI": "base_estado",
}

# Columnas que en el Excel están guardadas como fracción (0.0337 = 3.37%) y que
# queremos exportar ya multiplicadas por 100 para que el dashboard las pinte directo
# con un '%' al final, sin duplicar la lógica de formato en el HTML.
PCT_FRACTION_COLS = {"ctr", "hook", "kyc_por_install"}

# Mapeo de texto de ESTADO -> categoría normalizada + color de semáforo.
# Basado en los 12 estados reales que produce la fórmula ESTADO (que envuelve la fórmula
# vieja de Base CPI con dos gates nuevos: calidad de KYC e inactividad del anuncio):
# 🕐 En test / 🟡 En test / 🟢 Ganador temprano / 🟢 Escalar / 🔵 Mantener / 🟠 Iterar /
# 🟠 Revisar calidad (KYC bajo) / 🔴 Apagar / 🔴 Apagar ya / 🔴 Apagar (sin volumen) /
# ⚫ Ya apagado / ⚫ Off (revisar)
# IMPORTANTE: el orden de los checks importa (más específico primero), porque varias
# etiquetas comparten substrings (ej. "apagado" aparece dentro de "Ya apagado").
def clasificar_estado(estado_raw):
    if not estado_raw:
        return {"categoria": "sin_dato", "color": "gray", "label": "Sin dato"}
    s = str(estado_raw).lower()
    if "ya apagado" in s:
        return {"categoria": "apagado", "color": "graydark", "label": "Ya apagado"}
    if "off" in s:
        return {"categoria": "off_revisar", "color": "magenta", "label": "Off (revisar)"}
    if "revisar calidad" in s or "kyc bajo" in s:
        return {"categoria": "revisar_calidad", "color": "amber", "label": "Revisar calidad (KYC bajo)"}
    if "escal" in s or "ganador" in s:
        return {"categoria": "escalar", "color": "green", "label": str(estado_raw).strip()}
    if "apag" in s:
        return {"categoria": "apagar", "color": "red", "label": str(estado_raw).strip()}
    if "iter" in s:
        return {"categoria": "iterar", "color": "orange", "label": "Iterar"}
    if "manten" in s:
        return {"categoria": "mantener", "color": "blue", "label": "Mantener"}
    if "test" in s:
        return {"categoria": "en_test", "color": "gray", "label": "En test"}
    return {"categoria": "otro", "color": "gray", "label": str(estado_raw)}


def to_native(v):
    """Convierte tipos de openpyxl (date, datetime) a algo serializable en JSON."""
    if v is None:
        return None
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def col_map(ws, header_row=HEADER_ROW):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            m[str(v).strip()] = c
    return m


def read_benchmarks(wb):
    """Lee la hoja Benchmarks para poder explicar el semáforo con los umbrales reales
    (en vez de hardcodearlos en el HTML). Si la hoja no existe o cambia de layout,
    devuelve None y el dashboard cae a una explicación genérica."""
    if "Benchmarks" not in wb.sheetnames:
        return None
    bs = wb["Benchmarks"]
    try:
        rutas = []
        for r in range(5, 10):
            ruta = bs.cell(row=r, column=1).value
            cpi_ref = bs.cell(row=r, column=2).value
            if ruta:
                rutas.append({"ruta": ruta, "cpi_ref": to_native(cpi_ref)})
        config = {
            "mult_escalar": to_native(bs.cell(row=14, column=2).value),
            "mult_mantener": to_native(bs.cell(row=15, column=2).value),
            "mult_iterar": to_native(bs.cell(row=16, column=2).value),
            "ratio_apagar_ya": to_native(bs.cell(row=17, column=2).value),
            "min_installs": to_native(bs.cell(row=18, column=2).value),
            "ventana_dias": to_native(bs.cell(row=19, column=2).value),
            "gate_kyc_fraccion": to_native(bs.cell(row=21, column=2).value),
            "dias_sin_gasto_apagado": to_native(bs.cell(row=22, column=2).value),
            "fecha_corte": to_native(bs.cell(row=23, column=2).value),
            "min_kyc_gate": to_native(bs.cell(row=24, column=2).value),
        }
        return {"rutas": rutas, "config": config}
    except Exception:
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Ruta al Monitoreo_Creativos_Meta.xlsx ya recalculado")
    ap.add_argument("out_json", help="Ruta de salida, normalmente data.json")
    ap.add_argument("--fecha", default=None,
                     help='Etiqueta de fecha a mostrar en el dashboard, ej. "17 jul 2026". '
                          "Por defecto usa la fecha de hoy.")
    args = ap.parse_args()

    # data_only=True para leer VALORES calculados de las fórmulas (CPI, ESTADO, etc).
    # Esto requiere que el archivo haya sido recalculado con LibreOffice (recalc.py).
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: no encontré la hoja '{SHEET_NAME}'. Hojas disponibles: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET_NAME]
    cm = col_map(ws)

    faltantes = [h for h in COLS if h not in cm]
    if faltantes:
        print(f"AVISO: no encontré estas columnas en el encabezado (fila {HEADER_ROW}): "
              f"{faltantes}. Se omiten en el export.")

    piezas = []
    r = FIRST_DATA_ROW
    empty_streak = 0
    while empty_streak < 15 and r <= ws.max_row:
        pid_col = cm.get("Pieza ID")
        pid_val = ws.cell(row=r, column=pid_col).value if pid_col else None
        if pid_val is None or str(pid_val).strip() == "":
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0
        row = {}
        for header, key in COLS.items():
            if header not in cm:
                continue
            val = to_native(ws.cell(row=r, column=cm[header]).value)
            if key in PCT_FRACTION_COLS and isinstance(val, (int, float)):
                val = round(val * 100, 2)
            row[key] = val
        estado_info = clasificar_estado(row.get("estado"))
        row["estado_categoria"] = estado_info["categoria"]
        row["estado_color"] = estado_info["color"]
        row["estado_label"] = estado_info["label"]
        piezas.append(row)
        r += 1

    # Resumen agregado (lo recalculamos aquí en vez de parsear la hoja 'Resumen
    # Viernes' en texto libre, porque las categorías ya están en la columna ESTADO).
    resumen = {"escalar": 0, "mantener": 0, "iterar": 0, "revisar_calidad": 0, "apagar": 0,
               "en_test": 0, "apagado": 0, "off_revisar": 0, "sin_dato": 0, "otro": 0}
    gasto_total = 0
    installs_total = 0
    nuevas_q = 0
    for p in piezas:
        resumen[p["estado_categoria"]] = resumen.get(p["estado_categoria"], 0) + 1
        gasto_total += p.get("gasto") or 0
        installs_total += p.get("installs") or 0
        cohorte = str(p.get("cohorte") or "")
        if "nueva" in cohorte.lower() or "🆕" in cohorte:
            nuevas_q += 1

    # Top ganadores (escalar, ordenados por menor CPI vs ref = mejor eficiencia relativa)
    escalar = [p for p in piezas if p["estado_categoria"] == "escalar"]
    apagar = [p for p in piezas if p["estado_categoria"] == "apagar"]

    def cpi_vs_ref_num(p):
        v = p.get("cpi_vs_ref")
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    ganadores = sorted(escalar, key=cpi_vs_ref_num)[:5]
    perdedores = sorted(apagar, key=cpi_vs_ref_num, reverse=True)[:5]

    # Promedio real de KYC/install por ruta, solo entre piezas "confiables" (con KYC >= min_kyc_gate).
    # Es el número exacto contra el que la fórmula ESTADO compara a cada pieza para decidir si
    # aplica el gate de calidad. Lo recalculamos aquí (en vez de leerlo de una celda) porque en
    # el Excel es un AVERAGEIFS calculado por fila, no un valor único guardado en Benchmarks.
    bm = read_benchmarks(wb)
    kyc_prom_por_ruta = {}
    if bm and bm["config"].get("min_kyc_gate") is not None:
        min_kyc_gate = bm["config"]["min_kyc_gate"]
        por_ruta = {}
        for p in piezas:
            ruta = p.get("ruta")
            kyc_raw = p.get("kyc")
            kyc_ratio_pct = p.get("kyc_por_install")  # ya viene *100
            if ruta and kyc_raw is not None and kyc_ratio_pct is not None and kyc_raw >= min_kyc_gate:
                por_ruta.setdefault(ruta, []).append(kyc_ratio_pct)
        kyc_prom_por_ruta = {r: round(sum(vals) / len(vals), 2) for r, vals in por_ruta.items()}
    if bm is not None:
        bm["kyc_promedio_por_ruta"] = kyc_prom_por_ruta

    data = {
        "generado": args.fecha or date.today().strftime("%d %b %Y"),
        "total_piezas": len(piezas),
        "gasto_total": round(gasto_total, 2),
        "installs_total": installs_total,
        "nuevas_q": nuevas_q,
        "resumen": resumen,
        "benchmarks": bm,
        "ganadores": [{"pieza_id": p["pieza_id"], "ruta": p.get("ruta"),
                        "cpi": p.get("cpi"), "cpi_vs_ref": p.get("cpi_vs_ref")} for p in ganadores],
        "perdedores": [{"pieza_id": p["pieza_id"], "ruta": p.get("ruta"),
                         "cpi": p.get("cpi"), "cpi_vs_ref": p.get("cpi_vs_ref")} for p in perdedores],
        "piezas": piezas,
    }

    with open(args.out_json, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"OK: {len(piezas)} piezas exportadas a {args.out_json}")
    print(f"Resumen: {resumen}")


if __name__ == "__main__":
    main()
